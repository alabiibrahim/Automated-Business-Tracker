"""
Trucking Business Tracker – Excel Builder
==========================================
Generates a fully formatted 3-tab .xlsx workbook:
  Tab 1 – Revenue Tracker
  Tab 2 – Expense Tracker
  Tab 3 – Monthly Summary (auto-calculated via SUMPRODUCT formulas)

Requirements:
  pip install openpyxl

Run:
  python build_trucking_tracker.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
DARK_GRAY   = "404040"
YELLOW_HL   = "FFF2CC"
RED_LIGHT   = "FADBD8"
GREEN_LIGHT = "D5F5E3"
BORDER_CLR  = "BDC3C7"

# ── Reusable style helpers ─────────────────────────────────────────────────────
def full_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, color=WHITE):
    return Font(name="Arial", size=size, bold=True, color=color)

def body_font(bold=False):
    return Font(name="Arial", size=10, bold=bold, color=DARK_GRAY)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Helper: apply formatting to a data cell based on its column role ───────────
def style_data_cell(cell, col_idx, row_idx):
    bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
    cell.font = body_font()
    cell.fill = fill(bg)
    cell.border = full_border()
    if col_idx == 1:                        # Date column
        cell.number_format = "MM/DD/YYYY"
        cell.alignment = center()
    elif col_idx == 4:                      # Amount column
        cell.number_format = "$#,##0.00"
        cell.alignment = center()
    else:
        cell.alignment = left()


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 1 – REVENUE TRACKER
# ══════════════════════════════════════════════════════════════════════════════
ws_rev = wb.active
ws_rev.title = "💰 Revenue"
ws_rev.sheet_properties.tabColor = "17A589"
ws_rev.freeze_panes = "A3"           # Keep headers visible while scrolling

# --- Title banner (row 1) ---
ws_rev.merge_cells("A1:F1")
ws_rev["A1"] = "REVENUE TRACKER"
ws_rev["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws_rev["A1"].fill = fill(DARK_BLUE)
ws_rev["A1"].alignment = center()
ws_rev.row_dimensions[1].height = 36

# --- Column headers (row 2) ---
headers_rev   = ["Date", "Contractor Name", "Job / Load Description",
                 "Amount Paid ($)", "Payment Method", "Notes"]
col_widths_rev = [14,     24,               34,
                  18,              18,               30]

for col_idx, (h, w) in enumerate(zip(headers_rev, col_widths_rev), 1):
    cell = ws_rev.cell(row=2, column=col_idx, value=h)
    cell.font      = header_font()
    cell.fill      = fill(MED_BLUE)
    cell.alignment = center()
    cell.border    = full_border()
    ws_rev.column_dimensions[get_column_letter(col_idx)].width = w
ws_rev.row_dimensions[2].height = 28

# --- Sample data (rows 3–10) ---
sample_revenue = [
    ("2025-01-05", "ABC Freight Co.",   "Chicago → Detroit – 48ft flatbed",       2800,  "Check", ""),
    ("2025-01-12", "Midwest Logistics", "Detroit → Columbus – dry van",            1950,  "ACH",   ""),
    ("2025-01-20", "Smith Hauling LLC", "Columbus → Indianapolis – reefer",        2250,  "Check", "Fuel surcharge included"),
    ("2025-02-03", "ABC Freight Co.",   "Indianapolis → St. Louis",                2100,  "ACH",   ""),
    ("2025-02-14", "TruckLoad Direct",  "St. Louis → Kansas City",                 1600,  "Check", ""),
    ("2025-02-22", "Midwest Logistics", "Kansas City → Omaha",                     2400,  "ACH",   ""),
    ("2025-03-01", "Smith Hauling LLC", "Omaha → Denver – oversized",              3100,  "Check", "Oversize permit required"),
    ("2025-03-10", "ABC Freight Co.",   "Denver → Salt Lake City",                 2750,  "ACH",   ""),
]

for r_idx, row_data in enumerate(sample_revenue, 3):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_rev.cell(row=r_idx, column=c_idx, value=val)
        style_data_cell(cell, c_idx, r_idx)
    ws_rev.row_dimensions[r_idx].height = 20

# --- Empty entry rows (11–200) ---
for r_idx in range(11, 201):
    for c_idx in range(1, 7):
        style_data_cell(ws_rev.cell(row=r_idx, column=c_idx), c_idx, r_idx)
    ws_rev.row_dimensions[r_idx].height = 20

# --- Dropdown validation: Payment Method ---
dv_pay = DataValidation(
    type="list",
    formula1='"Check,ACH,Wire,Cash,Zelle,Other"',
    allow_blank=True,
    showDropDown=False      # False = show the arrow icon
)
dv_pay.sqref = "E3:E200"
ws_rev.add_data_validation(dv_pay)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 2 – EXPENSE TRACKER
# ══════════════════════════════════════════════════════════════════════════════
ws_exp = wb.create_sheet("🚛 Expenses")
ws_exp.sheet_properties.tabColor = "C0392B"
ws_exp.freeze_panes = "A3"

# --- Title banner ---
ws_exp.merge_cells("A1:F1")
ws_exp["A1"] = "EXPENSE TRACKER"
ws_exp["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws_exp["A1"].fill = fill(DARK_BLUE)
ws_exp["A1"].alignment = center()
ws_exp.row_dimensions[1].height = 36

# --- Column headers ---
headers_exp    = ["Date", "Expense Category", "Vendor / Description",
                  "Amount ($)", "Receipt #", "Notes"]
col_widths_exp = [14,     22,                 34,
                  14,           14,           30]

for col_idx, (h, w) in enumerate(zip(headers_exp, col_widths_exp), 1):
    cell = ws_exp.cell(row=2, column=col_idx, value=h)
    cell.font      = header_font()
    cell.fill      = fill("C0392B")   # Red theme for expenses
    cell.alignment = center()
    cell.border    = full_border()
    ws_exp.column_dimensions[get_column_letter(col_idx)].width = w
ws_exp.row_dimensions[2].height = 28

# --- Expense category list (used for dropdown AND summary formulas) ---
EXP_CATS = (
    "Fuel,Truck Payment,Insurance,Maintenance,Repairs,"
    "Permits & Licenses,Tolls,Tires,Oil & Fluids,Meals & Lodging,"
    "Dispatch Fees,Accounting & Legal,Phone,Miscellaneous"
)

# --- Sample data ---
sample_expenses = [
    ("2025-01-06", "Fuel",               "Pilot Flying J – Chicago",        320.45, "R001", ""),
    ("2025-01-07", "Fuel",               "Love's – Indianapolis",            298.20, "R002", ""),
    ("2025-01-10", "Truck Payment",      "First National Bank",             1450.00, "R003", "Monthly note"),
    ("2025-01-10", "Insurance",          "Progressive Commercial",           487.00, "R004", "Monthly premium"),
    ("2025-01-15", "Maintenance",        "Peterbilt Service Center",         215.00, "R005", "Oil change + filter"),
    ("2025-01-18", "Tolls",              "I-80 / I-94 EZPass",               47.60, "R006", ""),
    ("2025-01-25", "Permits & Licenses", "FMCSA Annual Update",               80.00, "R007", ""),
    ("2025-02-04", "Fuel",               "Pilot Flying J – St. Louis",       310.80, "R008", ""),
    ("2025-02-05", "Fuel",               "TA Travel Center – KC",            290.55, "R009", ""),
    ("2025-02-10", "Truck Payment",      "First National Bank",             1450.00, "R010", "Monthly note"),
    ("2025-02-10", "Insurance",          "Progressive Commercial",           487.00, "R011", "Monthly premium"),
    ("2025-02-18", "Repairs",            "Roadside Diesel Repair",           640.00, "R012", "Brake adjustment"),
    ("2025-02-20", "Tires",              "Pilot Tire Center",                885.00, "R013", "2 steer tires"),
    ("2025-03-02", "Fuel",               "Pilot Flying J – Omaha",           345.90, "R014", ""),
    ("2025-03-05", "Meals & Lodging",    "Holiday Inn Express",               89.00, "R015", "Denver layover"),
    ("2025-03-10", "Truck Payment",      "First National Bank",             1450.00, "R016", "Monthly note"),
    ("2025-03-10", "Insurance",          "Progressive Commercial",           487.00, "R017", "Monthly premium"),
    ("2025-03-10", "Permits & Licenses", "Colorado Oversize Permit",         175.00, "R018", "Oversized load Denver"),
]

for r_idx, row_data in enumerate(sample_expenses, 3):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws_exp.cell(row=r_idx, column=c_idx, value=val)
        style_data_cell(cell, c_idx, r_idx)
    ws_exp.row_dimensions[r_idx].height = 20

# --- Empty entry rows ---
for r_idx in range(21, 201):
    for c_idx in range(1, 7):
        style_data_cell(ws_exp.cell(row=r_idx, column=c_idx), c_idx, r_idx)
    ws_exp.row_dimensions[r_idx].height = 20

# --- Dropdown validation: Expense Category ---
dv_cat = DataValidation(
    type="list",
    formula1=f'"{EXP_CATS}"',
    allow_blank=True,
    showDropDown=False
)
dv_cat.sqref = "B3:B200"
ws_exp.add_data_validation(dv_cat)


# ══════════════════════════════════════════════════════════════════════════════
#  TAB 3 – MONTHLY SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws_sum = wb.create_sheet("📊 Monthly Summary")
ws_sum.sheet_properties.tabColor = "1F3864"
ws_sum.freeze_panes = "A4"

# --- Title banner ---
ws_sum.merge_cells("A1:I1")
ws_sum["A1"] = "MONTHLY FINANCIAL SUMMARY"
ws_sum["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
ws_sum["A1"].fill = fill(DARK_BLUE)
ws_sum["A1"].alignment = center()
ws_sum.row_dimensions[1].height = 36

# --- Subtitle hint ---
ws_sum.merge_cells("A2:I2")
ws_sum["A2"] = (
    "Formulas auto-calculate from the Revenue and Expenses tabs  •  "
    "Add new rows there and this summary updates instantly"
)
ws_sum["A2"].font = Font(name="Arial", size=9, italic=True, color="7F7F7F")
ws_sum["A2"].alignment = center()
ws_sum.row_dimensions[2].height = 18

# --- Section headers (row 3) ---
section_labels  = ["Month", "Total Revenue", "Total Expenses", "Net Profit",
                   "Fuel",  "Truck Payment", "Insurance", "Maintenance / Repairs", "Other Expenses"]
section_widths  = [14,       16,              16,           16,
                   14,        16,             14,           22,                     16]

for c_idx, (lbl, w) in enumerate(zip(section_labels, section_widths), 1):
    cell = ws_sum.cell(row=3, column=c_idx, value=lbl)
    cell.font      = header_font(size=10)
    cell.fill      = fill(DARK_BLUE) if c_idx <= 4 else fill("5D6D7E")
    cell.alignment = center()
    cell.border    = full_border()
    ws_sum.column_dimensions[get_column_letter(c_idx)].width = w
ws_sum.row_dimensions[3].height = 28

# --- Month definitions: label, first day, last day ---
months = [
    ("Jan 2025", "2025-01-01", "2025-01-31"),
    ("Feb 2025", "2025-02-01", "2025-02-28"),
    ("Mar 2025", "2025-03-01", "2025-03-31"),
    ("Apr 2025", "2025-04-01", "2025-04-30"),
    ("May 2025", "2025-05-01", "2025-05-31"),
    ("Jun 2025", "2025-06-01", "2025-06-30"),
    ("Jul 2025", "2025-07-01", "2025-07-31"),
    ("Aug 2025", "2025-08-01", "2025-08-31"),
    ("Sep 2025", "2025-09-01", "2025-09-30"),
    ("Oct 2025", "2025-10-01", "2025-10-31"),
    ("Nov 2025", "2025-11-01", "2025-11-30"),
    ("Dec 2025", "2025-12-01", "2025-12-31"),
]

# ── Formula builders ──────────────────────────────────────────────────────────
# SUMPRODUCT is used instead of SUMIFS so it works correctly in both
# Excel and Google Sheets (SUMIFS with date text comparisons can behave
# differently across platforms).

def rev_formula(start, end):
    """Sum all Revenue!D (Amount) where Revenue!A (Date) falls in [start, end]."""
    return (
        f"=SUMPRODUCT(('💰 Revenue'!D3:D200)*"
        f"('💰 Revenue'!A3:A200>=\"{start}\")*"
        f"('💰 Revenue'!A3:A200<=\"{end}\"))"
    )

def exp_formula(start, end, category=None):
    """
    Sum Expenses!D (Amount) within a date range, optionally filtered by category.

    category=None         → all expenses
    category="Fuel"       → single named category
    category="MAINT_REP"  → Maintenance OR Repairs combined
    category="OTHER"      → everything not in the named breakdown columns
    """
    base = (
        f"('🚛 Expenses'!D3:D200)*"
        f"('🚛 Expenses'!A3:A200>=\"{start}\")*"
        f"('🚛 Expenses'!A3:A200<=\"{end}\")"
    )
    if category is None:
        return f"=SUMPRODUCT({base})"

    elif category == "MAINT_REP":
        cat_filter = (
            f"*(('🚛 Expenses'!B3:B200=\"Maintenance\")"
            f"+('🚛 Expenses'!B3:B200=\"Repairs\")>0)"
        )
        return f"=SUMPRODUCT({base}{cat_filter})"

    elif category == "OTHER":
        # Exclude every category that has its own dedicated column
        excl = (
            f"*('🚛 Expenses'!B3:B200<>\"Fuel\")"
            f"*('🚛 Expenses'!B3:B200<>\"Truck Payment\")"
            f"*('🚛 Expenses'!B3:B200<>\"Insurance\")"
            f"*('🚛 Expenses'!B3:B200<>\"Maintenance\")"
            f"*('🚛 Expenses'!B3:B200<>\"Repairs\")"
        )
        return f"=SUMPRODUCT({base}{excl})"

    else:
        cat_filter = f"*('🚛 Expenses'!B3:B200=\"{category}\")"
        return f"=SUMPRODUCT({base}{cat_filter})"

# ── Build one row per month ────────────────────────────────────────────────────
for r_offset, (mth_lbl, start, end) in enumerate(months):
    row = 4 + r_offset
    bg  = WHITE if r_offset % 2 == 0 else LIGHT_GRAY

    def put(col, value, fmt="$#,##0.00", fnt=None, bg_color=None):
        c = ws_sum.cell(row=row, column=col, value=value)
        c.number_format = fmt
        c.font          = fnt or body_font()
        c.fill          = fill(bg_color or bg)
        c.alignment     = center()
        c.border        = full_border()

    # Col 1 – Month label
    c = ws_sum.cell(row=row, column=1, value=mth_lbl)
    c.font = body_font(bold=True); c.fill = fill(LIGHT_BLUE)
    c.alignment = center(); c.border = full_border()

    # Col 2 – Total Revenue  (green highlight)
    put(2, rev_formula(start, end),
        fnt=Font(name="Arial", size=10, bold=True, color="1A5276"),
        bg_color=GREEN_LIGHT)

    # Col 3 – Total Expenses  (red highlight)
    put(3, exp_formula(start, end),
        fnt=Font(name="Arial", size=10, bold=True, color="7B241C"),
        bg_color=RED_LIGHT)

    # Col 4 – Net Profit = Revenue − Expenses  (yellow highlight)
    b, d = get_column_letter(2), get_column_letter(3)
    put(4, f"={b}{row}-{d}{row}",
        fnt=Font(name="Arial", size=10, bold=True, color=DARK_GRAY),
        bg_color=YELLOW_HL)

    # Cols 5–9 – Expense breakdowns
    put(5, exp_formula(start, end, "Fuel"))
    put(6, exp_formula(start, end, "Truck Payment"))
    put(7, exp_formula(start, end, "Insurance"))
    put(8, exp_formula(start, end, "MAINT_REP"))
    put(9, exp_formula(start, end, "OTHER"))

    ws_sum.row_dimensions[row].height = 22

# ── Annual totals row ─────────────────────────────────────────────────────────
tot_row = 4 + len(months)
c = ws_sum.cell(row=tot_row, column=1, value="ANNUAL TOTAL")
c.font = header_font(size=10); c.fill = fill(DARK_BLUE)
c.alignment = center(); c.border = full_border()

for col in range(2, 10):
    col_ltr = get_column_letter(col)
    c = ws_sum.cell(row=tot_row, column=col,
                    value=f"=SUM({col_ltr}4:{col_ltr}{tot_row - 1})")
    c.number_format = "$#,##0.00"
    c.font          = header_font(size=10)
    c.fill          = fill(MED_BLUE)
    c.alignment     = center()
    c.border        = full_border()
ws_sum.row_dimensions[tot_row].height = 26


# ══════════════════════════════════════════════════════════════════════════════
#  SAVE
# ══════════════════════════════════════════════════════════════════════════════
OUTPUT = "TruckingBusiness_Tracker.xlsx"
wb.save(OUTPUT)
print(f"✅  Saved → {OUTPUT}")
